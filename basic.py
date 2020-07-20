from selenium import webdriver
from shutil import which
from selenium.webdriver.common.keys import Keys
chrome_path = which("chromedriver")
import pyautogui as pya
import pyperclip  # handy cross-platform clipboard text handler
import time
##import keyboard
import unittest
from openpyxl import load_workbook, cell
import openpyxl
import xlrd
import xlsxwriter




##new_workbook = xlsxwriter.Workbook()
##sheet = new_workbook.add_worksheet('stops')

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions


browser = webdriver.Chrome()
##"./chromedriver.exe"
driver = webdriver.Chrome(executable_path =chrome_path)

## reusable functions 
def copy_clipboard():
    pyperclip.copy("")  # <- This prevents last copy replacing current copy of null.
    pya.hotkey('ctrl', 'c')
    time.sleep(.1)  # ctrl-c is usually very fast but your program may execute faster
    return pyperclip.paste()

def function1(element1):
    for i in range(0,120):
        if(element1 != date_checkin):
            {
            driver.find_element_by_xpath("//span[@class='c-calendar-icon-next']").click()
            }
        element1 = driver.find_element_by_xpath('//*[@id="searchBoxCon"]/div/div/ul/li[2]/div/div[4]/div/div[1]/div[1]/h3').text             
        if(element1 == date_checkin):
            break


def scroll_fn():
    SCROLL_PAUSE_TIME = 5
    count = 0
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)
        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")      
        if new_height == last_height:
            try:
                
                check_elmnt = driver.find_element_by_xpath('//*[@id="ibu_hotel_container"]/div[1]/section/ul/div[4]/p/span')
                check_elmnt_bool=check_elmnt.is_enabled()
                if (check_elmnt_bool == True):
                    WebDriverWait(driver, 10).until(expected_conditions.element_to_be_clickable((By.XPATH, '//*[@id="ibu_hotel_container"]/div[1]/section/ul/div[4]/p/span'))).click()
                    
            ##if we encounetr no element the 
            except NoSuchElementException:
                count = count +1
                if(count ==2):
                    break

        last_height = new_height

## Inputs
city_name = 'Beijing'
date_checkin = 'Nov 2020'
adult_cnt = '2'
driver.get("https://www.trip.com")
driver.maximize_window

search_input = driver.find_element_by_id("hotels-destination")
search_input.send_keys(city_name)
time.sleep(5)
search_input.send_keys(Keys.TAB)

##select date 
element = driver.find_element_by_xpath('//*[@id="searchBoxCon"]/div/div/ul/li[2]/div/div[4]/div/div[1]/div[1]/h3').text
print('date selected is: ' +element)
function1(element)

##slect next calendar
driver.find_element_by_xpath("//li[@class='is-allow-hover'][contains(text(),'18')]").click()
driver.find_element_by_xpath("//li[@class='is-allow-hover'][contains(text(),'25')]").click()
driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "child-kid", " " )) and (((count(preceding-sibling::*) + 1) = 2) and parent::*)]//*[contains(concat( " ", @class, " " ), concat( " ", "u-icon-ic_plus", " " ))]').click()
##All inputs are enetered and search button is pressed
driver.find_element_by_xpath('//*[@id="searchBoxCon"]/div/div/ul/li[4]/div').click()
total_night = driver.find_element_by_xpath("//div[@class= 'nights']//span").text
print(total_night)


##print('done')
scroll_fn()
##whole page is scrolled down now

##All hotels stored in a list
all_hotels = driver.find_elements_by_class_name("list-card-title")
##print('rerached this part of code')

new_workbook = xlsxwriter.Workbook('dataset.xlsx')

sheet = new_workbook.add_worksheet('stops')
sheet.write(0, 0, "Hotel Name")
sheet.write(0, 1, "Hotel Star Rating")
sheet.write(0, 2, "Room Name")
sheet.write(0, 3, "Beds & Amenities")
sheet.write(0, 4, "Price per room")
sheet.write(0, 5, "No Of Adults")
sheet.write(0, 6, "No of days")
sheet.write(0, 7, "Date")
sheet.write(0, 8, "City")

##2 for loops 1st looping in each hotel as shown in output window other for number of rooms available in each room.

jump = 1
hotel_jump =1
for hotel in all_hotels:
    test = str(hotel_jump)
    ##WebDriverWait(driver, 10).until(expected_conditions.element_to_be_clickable((By.XPATH, var)).click()
    var = '(//div[@class = "list-card-title "]//span[@class = "name font-bold"])' + '[' + test + ']'
    test = driver.find_element_by_xpath(var)
   
    driver.execute_script("arguments[0].click();", test)
    time.sleep(2)
    window_after = driver.window_handles[1]
    driver.switch_to.window(window_after)
    ##Waiting for page to load
    WebDriverWait(driver,10).until(expected_conditions.element_to_be_clickable((By.XPATH, "//span[@class='mc-hd__login-btn']")))

    ###operations in new window
    try:
        title = driver.find_element_by_xpath("//section[contains(@class,'detail-baseinfo_title')]//h1").text
        print(title)
    except NoSuchElementException:
        pass
    ##add code for<if found exception add values such as NA>
    ## u-icon u-icon-star detail-baseinfo_title_level
    try:
        list_star = driver.find_elements_by_xpath("//i[@class='u-icon u-icon-diamond detail-baseinfo_title_level']")
        print(len(list_star))
        total_stars = len(list_star)
    except NoSuchElementException:
        pass
    

    toatl_rooms_hotel=driver.find_elements_by_xpath("//div[@class='roomname']")
    item_count = 1
    for each_room in toatl_rooms_hotel:
        
        create_xpath_room= '(//div[@class="roomname"])' + '[' + str(item_count) +']'        
        try:
            room_name=driver.find_element_by_xpath(create_xpath_room).text
            print(room_name)
        except NoSuchElementException:
            room_name = "Data Not Available"
            pass
    
        create_xpath_price = '(//div[contains(text(),"After tax ")])' + '[' + str(item_count) + ']'
        try:
            price = driver.find_element_by_xpath(create_xpath_price).text
            ##print thr original price
            print(price)
            str2 = str(price)
            moneystring = str2.split()
            print(moneystring[2])
        except NoSuchElementException:
            moneystring[2] = "Data not available"
            pass
   
        create_xpath_amenities = '(//*[@class = "u-icon u-icon-ic_new_fa_breakfast normal-icon"]//following-sibling::span)' + '[' + str(item_count) + ']'
        try:
            ammenities1 = driver.find_element_by_xpath(create_xpath_amenities).text
            print(ammenities1)
        except NoSuchElementException:
            ammenities1 = "Data not available"
            pass

        row_row = jump
        sheet.write(row_row, 0, title)
        sheet.write(row_row, 1, total_stars)
        sheet.write(row_row, 2, room_name)
        sheet.write(row_row, 3, ammenities1)
        sheet.write(row_row, 4, moneystring[2])

        sheet.write(row_row, 5, adult_cnt)
        sheet.write(row_row, 6, total_night)
        sheet.write(row_row, 7, date_checkin)
        sheet.write(row_row, 8, city_name)

        jump = jump+1
        item_count = item_count+1


    
    driver.close()
    window_before = driver.window_handles[0]
    driver.switch_to.window(window_before)
        ##wb.save("test-excel.xlsx")
    
    if(len(all_hotels) == hotel_jump):
        new_workbook.close()   
    
    hotel_jump = hotel_jump +1
    ##print('hotel jump value is :'+ str(hotel_jump))

    ##ammenities = driver.find_element_by_xpath("(//*[@class = 'u-icon u-icon-ic_new_fa_breakfast normal-icon']//following-sibling::span)[1]").get_attribute("innerHTML")

    

    
